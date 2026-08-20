# -*- coding: utf-8 -*-
"""Сборка docs/GPIO_ESP32.xlsx (EN) и docs/GPIO_ESP32.ru.xlsx (RU).

Оба файла собираются из ОДНОГО источника данных ниже — так они физически
не могут разойтись. Правится только этот скрипт, файлы пересобираются.
Печать: A4, альбомная ориентация.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

# Каталог docs/ рядом с этим скриптом (tools/ -> ..)
DOCS = os.path.join(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__))), 'docs')
FONT = 'Arial'

# ── Категории: ключ → (цвет, RU, EN) ─────────────────────────────
CATS = [
    ('power',  'FFF2CC', 'Питание',            'Power'),
    ('gnd',    'D9D9D9', 'Общий провод (GND)', 'Ground (GND)'),
    ('i2c',    'DEEAF6', 'Шина I²C',           'I²C bus'),
    ('sensor', 'E2EFDA', 'Датчик',             'Sensor'),
    ('relay',  'FCE4D6', 'Реле',               'Relay'),
    ('button', 'E4DFEC', 'Кнопка',             'Button'),
    ('led',    'FFF2CC', 'Индикация',          'Indication'),
    ('unused', 'F2F2F2', 'Не используется',    'Not used'),
]
CAT_COLOR = {k: c for k, c, _, _ in CATS}
CAT_TEXT = {'ru': {k: ru for k, _, ru, _ in CATS},
            'en': {k: en for k, _, _, en in CATS}}

DASH = '—'
NO = {'ru': 'нет', 'en': 'no'}
NO_INT_PU = {'ru': 'нет (внутр. INPUT_PULLUP)',
             'en': 'no (internal INPUT_PULLUP)'}
PU_10K = {'ru': '10 кОм → 3.3V (подтяжка вверх)',
          'en': '10 kΩ → 3.3 V (pull-up)'}
CAP_100N = {'ru': '100 нФ параллельно кнопке',
            'en': '100 nF across the button'}
BOTH_DASH = {'ru': DASH, 'en': DASH}
RELAY_CH = {'ru': 'Канал 8-канального блока реле',
            'en': 'Channel of the 8-ch relay board'}
TO_GND = {'ru': 'Замыкание на GND напрямую', 'en': 'Shorts straight to GND'}
FLASH_BUS = {'ru': 'Не используется (шина флеш-памяти)',
             'en': 'Not used (SPI flash bus)'}
GROUND = {'ru': 'Общий провод', 'en': 'Ground'}

# ── Данные: вывод платы, сигнал, категория, назначение, резистор,
#    конденсатор, примечание ───────────────────────────────────────
ROWS = [
 ('1', '3.3V', 'power',
  {'ru': 'Питание датчиков (выход)', 'en': 'Sensor supply (output)'},
  BOTH_DASH, BOTH_DASH,
  {'ru': 'Выход стабилизатора платы', 'en': 'Board regulator output'}),

 ('2', 'EN', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH,
  {'ru': 'Не используется', 'en': 'Not used'}),

 ('3', 'GPIO36', 'sensor',
  {'ru': 'YL-83 — датчик дождя, выход AO (аналоговый)',
   'en': 'YL-83 rain sensor, AO output (analog)'},
  NO, NO,
  {'ru': 'ADC1_CH0. Меньше значение = сильнее дождь. Пороги RAIN_ADC_WET / '
         'RAIN_ADC_DRY в config.h, подбираются по строке [RAIN] raw=…',
   'en': 'ADC1_CH0. Lower value = heavier rain. Thresholds RAIN_ADC_WET / '
         'RAIN_ADC_DRY live in config.h; tune them from the [RAIN] raw=… line'}),

 ('4', 'GPIO39', 'button',
  {'ru': 'Кнопка «Насос» (ручное вкл/выкл)',
   'en': 'Pump button (manual on/off)'},
  PU_10K, CAP_100N,
  {'ru': 'Второй контакт кнопки и конденсатора — на GND. Подтяжка '
         'ОБЯЗАТЕЛЬНА: у GPIO34-39 нет внутренней. Опрос, без прерывания '
         '(эррата GPIO36/39)',
   'en': 'Other side of button and capacitor goes to GND. Pull-up is '
         'MANDATORY: GPIO34-39 have none internally. Polled, never on an '
         'interrupt (GPIO36/39 erratum)'}),

 ('5', 'GPIO34', 'sensor',
  {'ru': 'AJ-SR04M — ECHO', 'en': 'AJ-SR04M — ECHO'},
  {'ru': 'Делитель 1,5 кОм + 2,2 кОм', 'en': 'Divider 1.5 kΩ + 2.2 kΩ'}, NO,
  {'ru': '1,5 кОм последовательно, 2,2 кОм с узла на GND (5V → ~2,65V)',
   'en': '1.5 kΩ in series, 2.2 kΩ from the node to GND (5 V → ~2.65 V)'}),

 ('6', 'GPIO35', 'button',
  {'ru': 'Кнопка «Увлажнитель» (ручное вкл/выкл)',
   'en': 'Humidifier button (manual on/off)'},
  PU_10K, CAP_100N,
  {'ru': 'Второй контакт кнопки и конденсатора — на GND. Подтяжка '
         'ОБЯЗАТЕЛЬНА: у GPIO34-39 нет внутренней',
   'en': 'Other side of button and capacitor goes to GND. Pull-up is '
         'MANDATORY: GPIO34-39 have none internally'}),

 ('7', 'GPIO32', 'relay',
  {'ru': 'Реле — Форточка ВЕРХ, открыть', 'en': 'Relay — upper vent, open'},
  NO, NO, RELAY_CH),
 ('8', 'GPIO33', 'relay',
  {'ru': 'Реле — Форточка НИЗ, открыть', 'en': 'Relay — lower vent, open'},
  NO, NO, RELAY_CH),
 ('9', 'GPIO25', 'relay',
  {'ru': 'Реле — Полив-1', 'en': 'Relay — irrigation 1'}, NO, NO, RELAY_CH),
 ('10', 'GPIO26', 'relay',
  {'ru': 'Реле — Полив-2', 'en': 'Relay — irrigation 2'}, NO, NO, RELAY_CH),
 ('11', 'GPIO27', 'relay',
  {'ru': 'Реле — Полив-3', 'en': 'Relay — irrigation 3'}, NO, NO, RELAY_CH),
 ('12', 'GPIO14', 'relay',
  {'ru': 'Реле — Увлажнитель', 'en': 'Relay — humidifier'}, NO, NO, RELAY_CH),

 ('13', 'GPIO12', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH,
  {'ru': 'НЕ ИСПОЛЬЗОВАТЬ. Strapping-пин выбора напряжения флеш-памяти: '
         'высокий уровень при старте переводит флеш в режим 1,8 В, что для '
         '3,3-вольтовой памяти платы означает отказ загрузки или повреждение',
   'en': 'DO NOT USE. Strapping pin selecting flash voltage: held high at '
         "reset it switches flash to 1.8 V mode, which on this board's "
         '3.3 V flash means a failed boot or a damaged chip'}),

 ('14', 'GND', 'gnd', GROUND, BOTH_DASH, BOTH_DASH, BOTH_DASH),

 ('15', 'GPIO13', 'sensor',
  {'ru': 'DHT22 — DATA (температура/влажность СНАРУЖИ)',
   'en': 'DHT22 — DATA (OUTDOOR temperature/humidity)'},
  NO, NO,
  {'ru': 'Только мониторинг. Автоматика на него не опирается: подменять им '
         'внутренний BME280 физически неверно',
   'en': 'Monitoring only. Control never falls back to it: substituting it '
         'for the indoor BME280 is physically wrong'}),

 ('16', 'GPIO9',  'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH, FLASH_BUS),
 ('17', 'GPIO10', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH, FLASH_BUS),
 ('18', 'GPIO11', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH, FLASH_BUS),

 ('19', 'Vin 5V', 'power',
  {'ru': 'Питание 5V', 'en': '5 V supply'}, BOTH_DASH, BOTH_DASH,
  {'ru': 'Вход от внешнего БП', 'en': 'Input from the external PSU'}),

 ('38', 'GND', 'gnd', GROUND, BOTH_DASH, BOTH_DASH, BOTH_DASH),

 ('37', 'GPIO23', 'button',
  {'ru': 'Кнопка «Форточки — открыть» (удержание)',
   'en': 'Vents OPEN button (hold)'},
  NO_INT_PU, NO, TO_GND),

 ('36', 'GPIO22', 'i2c',
  {'ru': 'I²C SCL (BME280 + GY-30)', 'en': 'I²C SCL (BME280 + GY-30)'},
  NO, NO, {'ru': 'Общая шина', 'en': 'Shared bus'}),

 ('35', 'GPIO1', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH,
  {'ru': 'Не используется (TX0)', 'en': 'Not used (TX0)'}),
 ('34', 'GPIO3', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH,
  {'ru': 'Не используется (RX0)', 'en': 'Not used (RX0)'}),

 ('33', 'GPIO21', 'i2c',
  {'ru': 'I²C SDA (BME280 + GY-30)', 'en': 'I²C SDA (BME280 + GY-30)'},
  NO, NO, {'ru': 'Общая шина', 'en': 'Shared bus'}),

 ('32', 'GND', 'gnd', GROUND, BOTH_DASH, BOTH_DASH, BOTH_DASH),

 ('31', 'GPIO19', 'button',
  {'ru': 'Кнопка «Форточки — закрыть» (удержание)',
   'en': 'Vents CLOSE button (hold)'},
  NO_INT_PU, NO, TO_GND),
 ('30', 'GPIO18', 'button',
  {'ru': 'Кнопка «Полив-1»', 'en': 'Irrigation 1 button'},
  NO_INT_PU, NO, TO_GND),

 ('29', 'GPIO5', 'relay',
  {'ru': 'Реле насоса скважины (SRD-05VDC-SL-C)',
   'en': 'Well pump relay (SRD-05VDC-SL-C)'},
  NO, NO,
  {'ru': 'Отдельное одноканальное реле, вне блока 8ch. GPIO5 — strapping-пин, '
         'но безобидный: влияет только на вывод отладочных сообщений загрузчика',
   'en': 'Separate single-channel relay, outside the 8-ch board. GPIO5 is a '
         'strapping pin but a harmless one: it only affects bootloader debug '
         'output'}),

 ('28', 'GPIO17', 'button',
  {'ru': 'Кнопка «Полив-2»', 'en': 'Irrigation 2 button'},
  NO_INT_PU, NO, TO_GND),
 ('27', 'GPIO16', 'button',
  {'ru': 'Кнопка «Полив-3»', 'en': 'Irrigation 3 button'},
  NO_INT_PU, NO, TO_GND),

 ('26', 'GPIO4', 'relay',
  {'ru': 'Реле — Форточка ВЕРХ, закрыть', 'en': 'Relay — upper vent, close'},
  NO, NO, RELAY_CH),

 ('25', 'GPIO0', 'sensor',
  {'ru': 'AJ-SR04M — TRIG', 'en': 'AJ-SR04M — TRIG'},
  PU_10K, NO,
  {'ru': 'Strapping-пин: низкий уровень при старте — режим прошивки. '
         'Подтяжка обеспечивает нормальную загрузку',
   'en': 'Strapping pin: held low at reset it enters flashing mode. The '
         'pull-up keeps normal boot'}),

 ('24', 'GPIO2', 'led',
  {'ru': 'LED heartbeat', 'en': 'Heartbeat LED'},
  {'ru': '4,7 кОм последовательно', 'en': '4.7 kΩ in series'}, NO,
  {'ru': 'Ограничительный резистор, не подтяжка',
   'en': 'Current-limiting resistor, not a pull-up'}),

 ('23', 'GPIO15', 'relay',
  {'ru': 'Реле — Форточка НИЗ, закрыть (вход IN7 релейного блока)',
   'en': 'Relay — lower vent, close (IN7 of the relay board)'},
  {'ru': '10 кОм → GND (подтяжка вниз)', 'en': '10 kΩ → GND (pull-down)'}, NO,
  {'ru': 'Защита от ложного включения реле при загрузке платы (GPIO15 — '
         'strapping-пин)',
   'en': 'Stops the relay from firing while the board boots (GPIO15 is a '
         'strapping pin)'}),

 ('22', 'GPIO8', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH, FLASH_BUS),
 ('21', 'GPIO7', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH, FLASH_BUS),
 ('20', 'GPIO6', 'unused', BOTH_DASH, BOTH_DASH, BOTH_DASH, FLASH_BUS),
]

# ── Подписи интерфейса ───────────────────────────────────────────
L = {
 'ru': {
  'sheet': 'GPIO распиновка',
  'title': 'SMART GREENHOUSE — РАСПИНОВКА ESP32 DOIT DEVKIT V1',
  'sub': 'Версия прошивки 6.2. Резисторы и конденсаторы указаны явно там, '
         'где они обязательны. Пины GPIO34–39 — только вход, внутренней '
         'подтяжки не имеют.',
  'head': ('Вывод платы', 'GPIO / сигнал', 'Назначение', 'Резистор',
           'Конденс.', 'Примечание', 'Категория'),
  'legend': 'Легенда категорий',
  'totals': 'Итого',
  't_res': 'Пинов с внешним резистором:',
  't_cap': 'Пинов с конденсатором:',
  't_gpio': 'Всего задействованных GPIO:',
  't_btn': 'Кнопок:',
  't_rel': 'Реле:',
  'src': 'Источник ограничений по выводам: ESP-IDF Programming Guide, '
         'раздел «GPIO & RTC GPIO», и ECO_and_Workarounds_for_Bugs_in_ESP32.',
  'foot': 'SmartGreenhouse — распиновка ESP32 (v6.2)',
  'page': 'Стр. &P из &N',
  'nomask': 'нет*',
 },
 'en': {
  'sheet': 'GPIO pinout',
  'title': 'SMART GREENHOUSE — ESP32 DOIT DEVKIT V1 PINOUT',
  'sub': 'Firmware version 6.2. Resistors and capacitors are spelled out '
         'wherever they are mandatory. GPIO34–39 are input-only and have no '
         'internal pull resistors.',
  'head': ('Board pin', 'GPIO / signal', 'Purpose', 'Resistor',
           'Capacitor', 'Note', 'Category'),
  'legend': 'Category legend',
  'totals': 'Totals',
  't_res': 'Pins with an external resistor:',
  't_cap': 'Pins with a capacitor:',
  't_gpio': 'GPIOs in use:',
  't_btn': 'Buttons:',
  't_rel': 'Relays:',
  'src': 'Source for the pin restrictions: ESP-IDF Programming Guide, '
         '"GPIO & RTC GPIO", and ECO_and_Workarounds_for_Bugs_in_ESP32.',
  'foot': 'SmartGreenhouse — ESP32 pinout (v6.2)',
  'page': 'Page &P of &N',
  'nomask': 'no*',
 },
}

thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WIDTHS = {'A': 6, 'B': 10, 'C': 26, 'D': 14, 'E': 12, 'F': 40, 'G': 13}


def build(lang, out_path):
    t = L[lang]
    wb = Workbook()
    ws = wb.active
    ws.title = t['sheet']

    ws['A1'] = t['title']
    ws['A1'].font = Font(name=FONT, size=14, bold=True)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    ws['A2'] = t['sub']
    ws['A2'].font = Font(name=FONT, size=9, italic=True, color='595959')
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
    ws.row_dimensions[2].height = 26

    HDR = 4
    for c, title in enumerate(t['head'], 1):
        cell = ws.cell(row=HDR, column=c, value=title)
        cell.font = Font(name=FONT, size=9, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='44546A')
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[HDR].height = 30

    r = HDR + 1
    for pin, sig, cat, purpose, res, cap, note in ROWS:
        vals = [pin, sig, purpose[lang], res[lang], cap[lang],
                note[lang], CAT_TEXT[lang][cat]]
        fill = PatternFill('solid', fgColor=CAT_COLOR[cat])
        warn = (sig == 'GPIO12')
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT, size=8,
                             bold=(warn or (c == 2 and cat != 'unused')),
                             color=('C00000' if warn else '000000'))
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal='center' if c in (1, 2, 4, 5) else 'left',
                vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 24
        r += 1
    d0, d1 = HDR + 1, r - 1

    r += 1
    ws.cell(row=r, column=1, value=t['legend']).font = Font(
        name=FONT, size=10, bold=True)
    r += 1
    legend_start = r
    for key, color, _ru, _en in CATS:
        cell = ws.cell(row=r, column=2, value=CAT_TEXT[lang][key])
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(name=FONT, size=8)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 15
        r += 1

    ws.cell(row=legend_start, column=4, value=t['src']).font = Font(
        name=FONT, size=8, italic=True, color='595959')

    r += 1
    ws.cell(row=r, column=1, value=t['totals']).font = Font(
        name=FONT, size=10, bold=True)
    r += 1

    # Шаблон «нет*» / «no*», а не точное сравнение: в колонке резисторов
    # есть значение «нет (внутр. INPUT_PULLUP)», и точное сравнение
    # засчитало бы пять кнопок как имеющие внешний резистор.
    nm = t['nomask']
    unused = CAT_TEXT[lang]['unused']
    power = CAT_TEXT[lang]['power']
    gnd = CAT_TEXT[lang]['gnd']
    button = CAT_TEXT[lang]['button']
    relay = CAT_TEXT[lang]['relay']
    totals = [
     (t['t_res'],  '=COUNTIFS(D{0}:D{1},"<>{2}",D{0}:D{1},"<>{3}")'
                   .format(d0, d1, nm, DASH)),
     (t['t_cap'],  '=COUNTIFS(E{0}:E{1},"<>{2}",E{0}:E{1},"<>{3}")'
                   .format(d0, d1, nm, DASH)),
     (t['t_gpio'], '=COUNTIF(G{0}:G{1},"<>{2}")-COUNTIF(G{0}:G{1},"{3}")'
                   '-COUNTIF(G{0}:G{1},"{4}")'
                   .format(d0, d1, unused, power, gnd)),
     (t['t_btn'],  '=COUNTIF(G{0}:G{1},"{2}")'.format(d0, d1, button)),
     (t['t_rel'],  '=COUNTIF(G{0}:G{1},"{2}")'.format(d0, d1, relay)),
    ]
    for label, formula in totals:
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=9)
        c = ws.cell(row=r, column=3, value=formula)
        c.font = Font(name=FONT, size=9, bold=True)
        c.alignment = Alignment(horizontal='left')
        ws.row_dimensions[r].height = 14
        r += 1

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    # Печать: A4 альбомная. Сумма ширин 121 ед. ≈ 9,6" при печатной области
    # альбомной A4 в 10,9" — «вписать в ширину» ничего не сжимает.
    ws.page_setup.paperSize = ws.PAPERSIZE_A4      # 9
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5,
                                  header=0.2, footer=0.2)
    ws.print_title_rows = '%d:%d' % (HDR, HDR)
    ws.print_area = 'A1:G%d' % (r - 1)
    ws.oddFooter.left.text = t['foot']
    ws.oddFooter.left.size = 8
    ws.oddFooter.right.text = t['page']
    ws.oddFooter.right.size = 8

    ws.freeze_panes = 'A%d' % (HDR + 1)
    ws.auto_filter.ref = 'A%d:G%d' % (HDR, d1)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return d0, d1


if __name__ == '__main__':
    for lang, name in (('en', 'GPIO_ESP32.xlsx'), ('ru', 'GPIO_ESP32.ru.xlsx')):
        path = os.path.join(DOCS, name)
        d0, d1 = build(lang, path)
        print('%-22s %s  строк данных %d (%d-%d)'
              % (name, lang.upper(), len(ROWS), d0, d1))
